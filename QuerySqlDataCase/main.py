import os.path
import pickle
import shutil
import time

import openpyxl
from openpyxl.styles import PatternFill

from QuerySqlDataCase.DataInfo import CountryAllLange, CityAllLange, PoiAllLange, PtaddrAllLange
from QuerySqlDataCase.QueryScript import QueryCountry, QueryCity, QueryPoi, QueryPtaddr
from Utils import SqlServer, load_config, LogUtils, ObjUtils


def query_country(db: SqlServer, iso_country_code: list = None, language_code: list = None):
    """
    查询国家数据
    支持指定国家代码，不指定时则查询所有国家；
    支持指定语言，默认为本国语言，多语言时默认第一个语言为查询语言
    返回一个dict {"ISO":class CountryAllLange}
    """
    country_query_results = db.query(QueryCountry(iso_country_code=iso_country_code, language_code=language_code))
    # country_query_results = [ class QueryResults, class QueryResults, class QueryResults...]
    # 在这里循环判断查询结果，将其生成CountryPreLange和CountryAllLange
    LogUtils.info("Query Country:%s ,With language:%s, country count:%s" % (
        iso_country_code, language_code, len(country_query_results)))
    data = {}
    for result in country_query_results:
        if result.ISO_COUNTRY_CODE in data.keys():
            data[result.ISO_COUNTRY_CODE].load_lange(result)
            LogUtils.warning("ADD CountryAllLange's lange for iso country code:%s-%s" % (
                result.ISO_COUNTRY_CODE, result.NAME_LANGUAGE))
        else:
            data[result.ISO_COUNTRY_CODE] = CountryAllLange(result)
            LogUtils.debug(
                "NEW CountryAllLange for iso country code:%s-%s" % (result.ISO_COUNTRY_CODE, result.NAME_LANGUAGE))
    return data


def query_city_by_country(db: SqlServer, c_data: dict):
    """
    根据query_country查询的结果查询相应的city数据
    返回一个dict {"ISO":{"CITY_POI_ID":class CityAllLange,
                        "CITY_POI_ID":class CityAllLange}
                "ISO":{...}}
    """
    data = {}
    for _iso in c_data:
        data[_iso] = query_city(db, iso_country_code=_iso, language_code=c_data[_iso].lange_list)
    return data


def query_city(db: SqlServer, iso_country_code: str, language_code: list):
    """
    查询城市数据
    需要指定国家代码
    支持指定语言，默认为本国语言
    返回一个dict {"CITY_POI_ID":class CityAllLange}
    """
    data = {}
    city_query_results = db.query(QueryCity(iso_country_code=iso_country_code, language_code=language_code))
    if len(city_query_results) < 6:
        LogUtils.warning("Query City for country %s ,With language:%s, result count:%s" % (
            iso_country_code, language_code, len(city_query_results)))
    else:
        LogUtils.info("Query City for country %s ,With language:%s, result count:%s" % (
            iso_country_code, language_code, len(city_query_results)))
    for result in city_query_results:
        if result.NAMED_PLACE_ID in data.keys():
            data[result.NAMED_PLACE_ID].load_lange(result)
        else:
            data[result.NAMED_PLACE_ID] = CityAllLange(result)
    return data


def query_poi_by_city(db: SqlServer, c_data: dict):
    """
    根据query_city的查询结果进行poi查询
    需要指定国家代码，和城市代码
    支持指定语言，默认为本国语言
    返回一个dict{"ISO":{"CITY_POI_ID":{"POI_ID":class PoiAllLange,
                                        "POI_ID":class PoiAllLange},
                        "CITY_POI_ID":{"POI_ID":class PoiAllLange,
                                        "POI_ID":class PoiAllLange}
                "ISO":{...}}
    """
    data = {}
    for _iso in c_data.keys():
        data[_iso] = {}
        for _c_id in c_data[_iso].keys():
            data[_iso][_c_id] = query_poi(db,
                                          iso_country_code=_iso,
                                          admin_order=_c_id,
                                          language_code=c_data[_iso][_c_id].lange_list)
    return data


def query_poi(db: SqlServer, iso_country_code: str, admin_order: str, language_code: list):
    """
    根据传入的信息查询poi的信息
    必须传入iso_country_code
    必须传入admin_order
    必须传入language_code
    如果传入poi_id则查询相应的poi_id内容
    返回一个dict{"POI_ID":class PoiAllLange,
                "POI_ID":class PoiAllLange}
    """
    data = {}
    for language in language_code:
        poi_query_results = db.query(QueryPoi(iso_country_code=iso_country_code,
                                              order8_id=admin_order,
                                              poi_id=list(data.keys()),
                                              language=language))

        LogUtils.info("Query Poi for country %s ,city:%s ,With language:%s, result count:%s" % (
            iso_country_code, admin_order, language_code, len(poi_query_results)))
        for result in poi_query_results:
            if result.POI_ID in data.keys():
                data[result.POI_ID].load_lange(result)
            else:
                data[result.POI_ID] = PoiAllLange(result)

    return data


def query_ptaddr_by_city(db: SqlServer, c_data: dict):
    """
    返回一个dict{"ISO":{"CITY_POI_ID":{"EDGE_ID":class PtaddrAllLange,
                                        "EDGE_ID":class PtaddrAllLange},
                        "CITY_POI_ID":{"EDGE_ID":class PtaddrAllLange,
                                        "EDGE_ID":class PtaddrAllLange}
                "ISO":{...}}
    """
    data = {}
    for _iso in c_data.keys():
        data[_iso] = {}
        for _c_id in c_data[_iso].keys():
            data[_iso][_c_id] = query_ptaddr(db,
                                             iso_country_code=_iso,
                                             order8_id=_c_id,
                                             language_code=c_data[_iso][_c_id].lange_list)

    return data


def query_ptaddr(db: SqlServer, iso_country_code: str, order8_id: str, language_code: list):
    """
    根据传入的数据查询ptaddr信息
    """
    data = {}
    for language in language_code:
        ptaddr_query_results = db.query(QueryPtaddr(iso_code=iso_country_code,
                                                    order_8=order8_id,
                                                    edge_id=list(data.keys()),
                                                    language=language))
        LogUtils.info("Query Ptaddr for country %s ,city:%s ,With language:%s, result count:%s" % (
            iso_country_code, order8_id, language_code, len(ptaddr_query_results)))
        for result in ptaddr_query_results:
            if result.EdgeID in data.keys():
                data[result.EdgeID].load_lange(result)
            else:
                data[result.EdgeID] = PtaddrAllLange(result)
    return data


def write_excel(save_path: str, here_db_name: str, country_data, city_data=None, poi_data=None, ptaddr_data=None):
    """
    :param save_path:
    :param country_data: 是CountryInfos类的列表
    :param city_data:
    :param poi_data:
    :param ptaddr_data:
    :return:
    """
    # 表格配色
    fill_country_header = PatternFill('solid', fgColor="76923C")
    fill_country_body = PatternFill('solid', fgColor="D7E3BC")
    fill_city_header = PatternFill('solid', fgColor="5F497A")
    fill_city_body = PatternFill('solid', fgColor="CCC1D9")
    fill_poi_header = PatternFill('solid', fgColor="31859B")
    fill_poi_body = PatternFill('solid', fgColor="C5DDE8")
    fill_ptaddr_header = PatternFill('solid', fgColor="E36C09")
    fill_ptaddr_body = PatternFill('solid', fgColor="FBD5B5")
    # 新建工作簿与工作表
    LogUtils.info('[开始构建表格]')
    wb = openpyxl.Workbook()
    sheet_index = 0
    wb.create_sheet(index=sheet_index, title='country_data')
    cur_sheet = wb.worksheets[sheet_index]
    cur_row_index = 1
    cur_col_index = 1
    # 设置第一行标题
    for attr in QueryCountry.QueryResults.arg:
        cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_country_header
        cur_col_index += 1
    cur_row_index += 1
    for pre_country in country_data:
        # 处理第一个国家的数据
        for cur_language_data in country_data[pre_country].all_lange_obj:
            # 处理第一个语言
            cur_col_index = 1
            for attr in QueryCountry.QueryResults.arg:
                cur_sheet.cell(cur_row_index, cur_col_index, getattr(cur_language_data, attr)).fill = fill_country_body
                cur_col_index += 1
            cur_row_index += 1
    # 新的标签页
    if city_data:
        sheet_index += 1
        wb.create_sheet(index=sheet_index, title='city_data')
        cur_sheet = wb.worksheets[sheet_index]
        cur_row_index = 1
        cur_col_index = 1
        # 设置第一行标题
        for attr in QueryCountry.QueryResults.arg:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_country_header
            cur_col_index += 1
        for attr in QueryCity.QueryResults.arg:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_city_header
            cur_col_index += 1
        cur_row_index += 1
        # 填数据
        for pre_country_code in city_data:
            for pre_city_code in city_data[pre_country_code]:
                for pre_city_lang in city_data[pre_country_code][pre_city_code].all_lange_obj:
                    cur_col_index = 1
                    for attr in QueryCountry.QueryResults.arg:
                        cur_sheet.cell(cur_row_index, cur_col_index,
                                       getattr(country_data[pre_country_code].get_lange(pre_city_lang.NAME_LANGUAGE),
                                               attr)).fill = fill_country_body
                        cur_col_index += 1
                    for attr in QueryCity.QueryResults.arg:
                        cur_sheet.cell(cur_row_index, cur_col_index, getattr(pre_city_lang, attr)).fill = fill_city_body
                        cur_col_index += 1
                    cur_row_index += 1
    if poi_data:
        poi_page_country_col = ['ISO_COUNTRY_CODE', 'COUNTRY_ID', 'NAME', 'CAPITAL', 'LANGUAGE_CODE']
        poi_page_city_col = ['CITY_NAME', 'CITY_POI_ID', 'NAMED_PLACE_ID']
        sheet_index += 1
        wb.create_sheet(index=sheet_index, title='poi_data')
        cur_sheet = wb.worksheets[sheet_index]
        cur_row_index = 1
        cur_col_index = 1
        # 设置第一行标题
        # 设置国家列
        for attr in poi_page_country_col:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_country_header
            cur_col_index += 1
        # 设置城市列
        for attr in poi_page_city_col:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_city_header
            cur_col_index += 1
        # 设置poi列
        for attr in QueryPoi.QueryResults.arg:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_poi_header
            cur_col_index += 1
        cur_row_index += 1
        # 填数据
        for pre_country_code in poi_data:
            for pre_city_code in poi_data[pre_country_code]:
                for pre_poi_code in poi_data[pre_country_code][pre_city_code]:
                    for pre_poi_lang in poi_data[pre_country_code][pre_city_code][pre_poi_code].all_lange_obj:
                        cur_col_index = 1
                        for attr in poi_page_country_col:
                            cur_sheet.cell(cur_row_index, cur_col_index, getattr(
                                country_data[pre_country_code].get_lange(pre_poi_lang.LANGUAGE_POI),
                                attr)).fill = fill_country_body
                            cur_col_index += 1
                        for attr in poi_page_city_col:
                            cur_sheet.cell(cur_row_index, cur_col_index,
                                           getattr(city_data[pre_country_code][pre_city_code].get_lange(
                                               pre_poi_lang.LANGUAGE_POI), attr)).fill = fill_city_body
                            cur_col_index += 1
                        for attr in QueryPoi.QueryResults.arg:
                            cur_sheet.cell(cur_row_index, cur_col_index,
                                           getattr(pre_poi_lang, attr)).fill = fill_poi_body
                            cur_col_index += 1
                        cur_row_index += 1
    # ptaddr分页
    if ptaddr_data:
        ptaddr_page_country_col = ['ISO_COUNTRY_CODE', 'COUNTRY_ID', 'NAME', 'CAPITAL', 'LANGUAGE_CODE']
        ptaddr_page_city_col = ['CITY_NAME', 'CITY_POI_ID', 'NAMED_PLACE_ID']
        sheet_index += 1
        wb.create_sheet(index=sheet_index, title='ptaddr_data')
        cur_sheet = wb.worksheets[sheet_index]
        cur_row_index = 1
        cur_col_index = 1
        # 设置第一行标题
        # 设置国家列
        for attr in ptaddr_page_country_col:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_country_header
            cur_col_index += 1
        # 设置城市列
        for attr in ptaddr_page_city_col:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_city_header
            cur_col_index += 1
        for attr in QueryPtaddr.QueryResults.arg:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_ptaddr_header
            cur_col_index += 1
        cur_row_index += 1
        # 填数据
        for pre_country_code in ptaddr_data:
            for pre_city_code in ptaddr_data[pre_country_code]:
                for pre_ptaddr_code in ptaddr_data[pre_country_code][pre_city_code]:
                    for pre_ptaddr_lange in ptaddr_data[pre_country_code][pre_city_code][pre_ptaddr_code].all_lange_obj:
                        cur_col_index = 1
                        for attr in ptaddr_page_country_col:
                            cur_sheet.cell(cur_row_index, cur_col_index, getattr(
                                country_data[pre_country_code].get_lange(pre_ptaddr_lange.Road_language),
                                attr)).fill = fill_country_body
                            cur_col_index += 1
                        for attr in ptaddr_page_city_col:
                            cur_sheet.cell(cur_row_index, cur_col_index,
                                           getattr(city_data[pre_country_code][pre_city_code].get_lange(
                                               pre_ptaddr_lange.Road_language), attr)).fill = fill_city_body
                            cur_col_index += 1
                        for attr in QueryPtaddr.QueryResults.arg:
                            cur_sheet.cell(cur_row_index, cur_col_index,
                                           getattr(pre_ptaddr_lange, attr)).fill = fill_ptaddr_body
                            cur_col_index += 1
                        cur_row_index += 1
    # 保存工作表

    wb.save(f'{save_path}/{here_db_name}.xlsx')
    LogUtils.info('[表格构建结束]')


def main(sql_db_name: str, iso_country_code: list = None, language_code: list = None):
    time_now = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time()))
    work_dir = os.path.join(".", f"{sql_db_name}_{time_now}")
    if not os.path.exists(os.path.join(".", work_dir)):
        os.mkdir(os.path.join(".", work_dir))
        LogUtils.info("Mkdir: %s" % work_dir)
    dbs = SqlServer(**load_config().get("sqlserver"), database=sql_db_name)
    LogUtils.info("Connect to sql server: %s" % load_config().get("sqlserver"))
    country_data = query_country(dbs, iso_country_code=iso_country_code, language_code=language_code)
    ObjUtils.save_obj(country_data, work_dir, "country_data")
    city_data = query_city_by_country(dbs, country_data)
    ObjUtils.save_obj(city_data, work_dir, "city_data")
    poi_data = query_poi_by_city(dbs, city_data)
    ObjUtils.save_obj(poi_data, work_dir, "poi_data")
    ptaddr_data = query_ptaddr_by_city(dbs, city_data)
    ObjUtils.save_obj(ptaddr_data, work_dir, "ptaddr_data")
    dbs.close()
    write_excel(save_path=f'./{work_dir}', here_db_name=sql_db_name, country_data=country_data, city_data=city_data,
                poi_data=poi_data, ptaddr_data=ptaddr_data)
    shutil.copy(os.path.join(".", "QueryDataCase.log"), f"./{work_dir}")


def debug(debug_dir: str):
    # 从文件中加载实例
    with open(f'./{debug_dir}/country_data.pkl', 'rb') as f:
        country_data = pickle.load(f)
    with open(f'./{debug_dir}/city_data.pkl', 'rb') as f:
        city_data = pickle.load(f)
    with open(f'./{debug_dir}/poi_data.pkl', 'rb') as f:
        poi_data = pickle.load(f)
    with open(f'./{debug_dir}/ptaddr_data.pkl', 'rb') as f:
        ptaddr_data = pickle.load(f)
    write_excel(save_path=f'./{debug_dir}_debug', here_db_name="debug", country_data=country_data,
                city_data=city_data, poi_data=poi_data, ptaddr_data=ptaddr_data)
    shutil.move(os.path.join(".", "QueryDateCase.log"), f"./{debug_dir}_debug")



if __name__ == '__main__':
    # main("HERE_MEA_S231R2", iso_country_code=["SAU"], language_code=["ARA", "ENG"])
    # main("HERE_EEU_S231R2", iso_country_code=["RUS"])
    main("HERE_EEU_S231R2")
