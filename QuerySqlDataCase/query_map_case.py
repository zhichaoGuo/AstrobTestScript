import time

import openpyxl
from openpyxl.styles import PatternFill

from DataInfo import CountryInfo, CityInfo, PoiInfo, PtaddrInfo, CountryInfos, CityInfos, PoiInfos, PtaddrInfos, \
    CountryAllLange
from QueryScript import QueryAllCountry, QueryAllCity, QueryCapital, QueryAllPoi, QueryAllPtaddr, QueryTotal, \
    QueryLevel, QueryCountry
from Utils import load_config, SqlServer, LogUtils


def query_all_country(db: SqlServer, iso_county_code: list = None, default_lang: str = 'ENG', mult_lang: list = []):
    ret = []
    all_country_res = db.query(QueryAllCountry(iso_country_code=iso_county_code, name_language=default_lang))
    # 开始处理多语言并转换为数据类：
    for country in all_country_res:
        capital = db.query(QueryCapital(country.ISO_COUNTRY_CODE, country.COUNTRY_ID, country.NAME_LANGUAGE))
        total = db.query(QueryTotal(country.ISO_COUNTRY_CODE))
        level = db.query(QueryLevel(country.ISO_COUNTRY_CODE))
        cur_country = CountryInfo(country, capital, total, level)
        cur_country_mult_lang = CountryInfos(cur_country)
        for lang in mult_lang:
            mult_country = db.query(QueryAllCountry(iso_country_code=[country.ISO_COUNTRY_CODE], name_language=lang))
            if mult_country:
                capital = db.query(QueryCapital(country.ISO_COUNTRY_CODE, country.COUNTRY_ID, lang))
                cur_country = CountryInfo(mult_country[0], capital, total, level)
                cur_country_mult_lang.load_other_language(cur_country)
        ret.append(cur_country_mult_lang)
    return ret





def query_all_city(db: SqlServer, country: CountryInfo, default_language: str = 'ENG', mult_language: list = []):
    ret = []
    all_city_res = db.query(QueryAllCity(country.ISO_COUNTRY_CODE, poi_language=default_language))
    # 开始处理多语言并转换为数据类
    for city in all_city_res:
        cur_city = CityInfo(city)
        cur_city_mult_lang = CityInfos(cur_city)
        for lang in mult_language:
            mult_city = db.query(
                QueryAllCity(country.ISO_COUNTRY_CODE, city_poi_id=cur_city.POI_ID, poi_language=lang))
            if mult_city:
                cur_city = CityInfo(mult_city[0])
                cur_city_mult_lang.load_other_language(cur_city)
        ret.append(cur_city_mult_lang)
    return ret


def query_all_poi(db: SqlServer, country: CountryInfo, city: CityInfo, default_language='ENG',
                  mult_language: list = []):
    ret = []
    all_poi_res = db.query(QueryAllPoi(country.ISO_COUNTRY_CODE, city.NAMED_PLACE_ID, language=default_language))
    for poi in all_poi_res:
        cur_poi = PoiInfo(poi)
        cur_poi_mult_lang = PoiInfos(cur_poi)
        for lang in mult_language:
            mult_poi = db.query(
                QueryAllPoi(country.ISO_COUNTRY_CODE, city.NAMED_PLACE_ID, poi_id=cur_poi.POI_ID, language=lang))
            if mult_poi:
                cur_poi = PoiInfo(mult_poi[0])
                cur_poi_mult_lang.load_other_language(cur_poi)
        ret.append(cur_poi_mult_lang)
    return ret


def query_all_patddr(db: SqlServer, country: CountryInfo, city: CityInfo, default_language: str = 'ENG',
                     mult_language: list = []):
    ret = []
    all_ptaddr_res = db.query(
        QueryAllPtaddr(country.ISO_COUNTRY_CODE, country.COUNTRY_ID, city.NAMED_PLACE_ID, language=default_language))
    for ptaddr in all_ptaddr_res:
        cur_ptaddr = PtaddrInfo(ptaddr)
        cur_ptaddr_mult_lang = PtaddrInfos(cur_ptaddr)
        for lang in mult_language:
            mult_ptaddr = db.query(QueryAllPtaddr(country.ISO_COUNTRY_CODE, country.COUNTRY_ID, city.NAMED_PLACE_ID,
                                                  ptaddr_id=cur_ptaddr.ADDRESS_POINT_ID, language=lang))
            if mult_ptaddr:
                cur_ptaddr = PtaddrInfo(mult_ptaddr[0])
                cur_ptaddr_mult_lang.load_other_language(cur_ptaddr)
        ret.append(cur_ptaddr_mult_lang)
    return ret


def write_excel(save_path: str, country_data, city_data=None, poi_data=None, ptaddr_data=None):
    """

    :param save_path:
    :param country_data: 是CountryInfos类的列表
    :param city_data:
    :param poi_data:
    :param ptaddr_data:
    :return:
    """
    # 表格配色
    fill_country_header = PatternFill('solid', fgColor="76923C")
    fill_country_body = PatternFill('solid', fgColor="D7E3BC")
    fill_city_header = PatternFill('solid', fgColor="5F497A")
    fill_city_body = PatternFill('solid', fgColor="CCC1D9")
    fill_poi_header = PatternFill('solid', fgColor="31859B")
    fill_poi_body = PatternFill('solid', fgColor="C5DDE8")
    fill_ptaddr_header = PatternFill('solid', fgColor="E36C09")
    fill_ptaddr_body = PatternFill('solid', fgColor="FBD5B5")
    # 新建工作簿与工作表
    LogUtils.info('[开始构建表格]')
    wb = openpyxl.Workbook()
    sheet_index = 0
    wb.create_sheet(index=sheet_index, title='country_data')
    cur_sheet = wb.worksheets[sheet_index]
    cur_row_index = 1
    cur_col_index = 1
    # 设置第一行标题
    for attr in CountryInfo.AllAttr:
        cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_country_header
        cur_col_index += 1
    cur_row_index += 1
    for cur_country_infos in country_data:
        # 处理第一个国家的数据
        for cur_language_data in cur_country_infos.data.values():
            # 处理第一个语言
            cur_col_index = 1
            for attr in CountryInfo.AllAttr:
                cur_sheet.cell(cur_row_index, cur_col_index, getattr(cur_language_data, attr)).fill = fill_country_body
                cur_col_index += 1
            cur_row_index += 1
    # 新的标签页
    if city_data:
        sheet_index += 1
        wb.create_sheet(index=sheet_index, title='city_data')
        cur_sheet = wb.worksheets[sheet_index]
        cur_row_index = 1
        cur_col_index = 1
        # 设置第一行标题
        for attr in CountryInfo.AllAttr:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_country_header
            cur_col_index += 1
        for attr in CityInfo.AllAttr:
            cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_city_header
            cur_col_index += 1
        cur_row_index += 1
        # 填数据
        for cur_country_infos in country_data:
            # 第一个国家
            for cur_city_infos in city_data[cur_country_infos.default_language_info.ISO_COUNTRY_CODE]:
                # 第一个城市
                for cur_city_info in cur_city_infos.data.values():
                    # 第一个语言
                    cur_col_index = 1
                    for attr in CountryInfo.AllAttr:
                        cur_sheet.cell(cur_row_index, cur_col_index,
                                       getattr(cur_country_infos.get_country_by_language(cur_city_info.LANGUAGE_CODE),
                                               attr)).fill = fill_country_body
                        cur_col_index += 1
                    for attr in CityInfo.AllAttr:
                        cur_sheet.cell(cur_row_index, cur_col_index, getattr(cur_city_info, attr)).fill = fill_city_body
                        cur_col_index += 1
                    cur_row_index += 1
        # poi分页
        if poi_data:
            poi_page_country_col = ['ISO_COUNTRY_CODE', 'COUNTRY_ID', 'NAME', 'CAPITAL', 'LANGUAGE']
            poi_page_city_col = ['CITY_NAME', 'POI_ID', 'NAMED_PLACE_ID']
            sheet_index += 1
            wb.create_sheet(index=sheet_index, title='poi_data')
            cur_sheet = wb.worksheets[sheet_index]
            cur_row_index = 1
            cur_col_index = 1
            # 设置第一行标题
            # 设置国家列
            for attr in poi_page_country_col:
                cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_country_header
                cur_col_index += 1
            # 设置城市列
            for attr in poi_page_city_col:
                cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_city_header
                cur_col_index += 1
            for attr in PoiInfo.AllAttr:
                cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_poi_header
                cur_col_index += 1
            cur_row_index += 1
            # 填数据
            for cur_country_infos in country_data:
                # 第一个国家
                for cur_city_infos in city_data[cur_country_infos.default_language_info.ISO_COUNTRY_CODE]:
                    # 第一个城市
                    for cur_poi_infos in poi_data[cur_city_infos.default_language_info.NAMED_PLACE_ID]:
                        # 第一个poi
                        for cur_poi_info in cur_poi_infos.data.values():
                            # 第一个语言
                            cur_col_index = 1
                            # 填国家信息
                            for attr in poi_page_country_col:
                                cur_sheet.cell(cur_row_index, cur_col_index,
                                               getattr(cur_country_infos.get_country_by_language(
                                                   cur_poi_info.LANGUAGE_POI),
                                                   attr)).fill = fill_country_body
                                cur_col_index += 1
                            # 填城市信息
                            for attr in poi_page_city_col:
                                cur_sheet.cell(cur_row_index, cur_col_index,
                                               getattr(cur_city_infos.get_city_by_language(cur_poi_info.LANGUAGE_POI),
                                                       attr)).fill = fill_city_body
                                cur_col_index += 1
                            # 填poi信息
                            for attr in PoiInfo.AllAttr:
                                cur_sheet.cell(cur_row_index, cur_col_index,
                                               getattr(cur_poi_info, attr)).fill = fill_poi_body
                                cur_col_index += 1
                            cur_row_index += 1
        # ptaddr分页
        if ptaddr_data:
            ptaddr_page_country_col = ['ISO_COUNTRY_CODE', 'COUNTRY_ID', 'NAME', 'CAPITAL', 'LANGUAGE']
            ptaddr_page_city_col = ['CITY_NAME', 'POI_ID', 'NAMED_PLACE_ID']
            sheet_index += 1
            wb.create_sheet(index=sheet_index, title='ptaddr_data')
            cur_sheet = wb.worksheets[sheet_index]
            cur_row_index = 1
            cur_col_index = 1
            # 设置第一行标题
            # 设置国家列
            for attr in ptaddr_page_country_col:
                cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_country_header
                cur_col_index += 1
            # 设置城市列
            for attr in ptaddr_page_city_col:
                cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_city_header
                cur_col_index += 1
            for attr in PtaddrInfo.AllAttr:
                cur_sheet.cell(cur_row_index, cur_col_index, attr).fill = fill_ptaddr_header
                cur_col_index += 1
            cur_row_index += 1
            # 填数据
            for cur_country_infos in country_data:
                # 第一个国家
                for cur_city_infos in city_data[cur_country_infos.default_language_info.ISO_COUNTRY_CODE]:
                    # 第一个城市
                    for cur_ptaddr_infos in ptaddr_data[cur_city_infos.default_language_info.NAMED_PLACE_ID]:
                        # 第一个ptaddr
                        for cur_ptaddr_info in cur_ptaddr_infos.data.values():
                            # 第一个语言
                            cur_col_index = 1
                            # 填国家信息
                            for attr in ptaddr_page_country_col:
                                cur_sheet.cell(cur_row_index, cur_col_index,
                                               getattr(cur_country_infos.get_country_by_language(
                                                   cur_ptaddr_info.Road_language),
                                                   attr)).fill = fill_country_body
                                cur_col_index += 1
                            # 填城市信息
                            for attr in ptaddr_page_city_col:
                                cur_sheet.cell(cur_row_index, cur_col_index,
                                               getattr(
                                                   cur_city_infos.get_city_by_language(cur_ptaddr_info.Road_language),
                                                   attr)).fill = fill_city_body
                                cur_col_index += 1
                            # 填ptaddr信息
                            for attr in PtaddrInfo.AllAttr:
                                cur_sheet.cell(cur_row_index, cur_col_index,
                                               getattr(cur_ptaddr_info, attr)).fill = fill_ptaddr_body
                                cur_col_index += 1
                            cur_row_index += 1
    # 保存工作表
    db_name = load_config().get("sqlserver").get("database")
    wb.save(f'{save_path}/{db_name}.xlsx')
    LogUtils.info('[表格构建结束]')


def query_map_info(iso_country_code: list = None, default_lang: str = 'ENG', mult_lang: list = []):
    """
    查询并抽取地图数据用例并输出excel
    :param iso_country_code: 需要查询的国家iso编码，默认为空则查询数据库内所有国家
    :param default_lang: 查询的主要语言，默认为英语，后续查询都依赖于这个语言
    :param mult_lang: 查询的多语言，默认为空则不查询多语言
    :return:
    """
    db = SqlServer(**load_config().get("sqlserver"))
    all_country_info = query_all_country(db,
                                         iso_county_code=iso_country_code,
                                         default_lang=default_lang,
                                         mult_lang=mult_lang)
    all_city_info = {}
    all_poi_info = {}
    all_ptaddr_info = {}
    for country_infos in all_country_info:
        cur_country_code = country_infos.default_language_info.ISO_COUNTRY_CODE
        all_city_info[cur_country_code] = query_all_city(db,
                                                         country_infos.default_language_info,
                                                         default_language=default_lang,
                                                         mult_language=mult_lang)
        for city_infos in all_city_info[cur_country_code]:
            cur_city_id = city_infos.default_language_info.NAMED_PLACE_ID
            all_poi_info[cur_city_id] = query_all_poi(db,
                                                      country_infos.default_language_info,
                                                      city=city_infos.default_language_info,
                                                      default_language=default_lang,
                                                      mult_language=mult_lang)
            all_ptaddr_info[cur_city_id] = query_all_patddr(db,
                                                            country_infos.default_language_info,
                                                            city=city_infos.default_language_info,
                                                            default_language=default_lang,
                                                            mult_language=mult_lang)
    db.close()
    time.sleep(1)
    write_excel('.', country_data=all_country_info, city_data=all_city_info, poi_data=all_poi_info,
                ptaddr_data=all_ptaddr_info)


if __name__ == '__main__':
    # query_map_info(iso_country_code=['RUS'], default_lang='RUS')
    db = SqlServer(**load_config().get("sqlserver"))
    query_country(db, iso_country_code=['RUS', 'POL'], language_code=['ENG', 'CHI'])
    db.close()
