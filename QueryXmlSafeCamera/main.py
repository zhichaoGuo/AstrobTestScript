import os
import pickle
import random
import shutil
import time
from typing import Union
from xml.etree.ElementTree import Element

import openpyxl

from QueryXmlEVCase.main import return_all_folder_without_tag, unzip_all_xml, return_all_node_in_xml
from QueryXmlSafeCamera.SafeCamera import SafeCamera, CameraType
from Utils import LogUtils, PathUtils, ObjUtils

'''
WHEN 'redlight' THEN '1' 
WHEN 'Speed' THEN '2'
WHEN 'RedLightAndSpeed' THEN '3'
WHEN 'sectionstart' THEN '4'
WHEN 'sectionend' THEN '5'
WHEN 'buslane' THEN '6'
WHEN 'distance' THEN '7'

WHEN 'Permanent' THEN 'P'
WHEN 'Mobile' THEN 'M'
WHEN 'PreAnnounced' THEN 'A'
'''


def return_camera_type(ele: Element) -> Union[str, None, bool]:
    try:
        return ele.find('Details').find('Camera').find('CameraType').text.upper()
    except Exception as err:
        return False


def return_section_end_name(start_name: str) -> str:
    """
    start name : AD-0551661_S_1
    end name:AD-0551661_E_1
    """
    return start_name.replace('_S_', '_E_')


def get_section_end_node(aim_start_node: list, all_end_node: list) -> list:
    """
    此处可以优化逻辑减少耗时
    此处还可以加排序
    """
    out_list = []
    for node in aim_start_node:
        aim_name = return_section_end_name(node.find('Identity').find('POI_Entity_ID').text)
        find_flag = False
        for end_node in all_end_node:
            if end_node.find('Identity').find('POI_Entity_ID').text == aim_name:
                out_list.append(end_node)
                find_flag = True
                break
        if not find_flag:
            LogUtils.warning('section end node not find! id:%s' % aim_name)
    return out_list


def write_excel(count_data, data, save_name: str):
    wb = openpyxl.Workbook()
    # 创建总览
    sheet_index = 0
    wb.create_sheet(index=sheet_index, title='country count')
    work_sheet = wb.worksheets[0]
    cur_row_index = 1
    cur_col_index = 1
    for _title in ['CC', *CameraType.ALL_TYPE, 'ALL']:
        work_sheet.cell(cur_row_index, cur_col_index, _title)
        cur_col_index += 1
    cur_row_index += 1
    # 写数量统计
    for _country in count_data.keys():
        cur_col_index = 1
        work_sheet.cell(cur_row_index, cur_col_index, _country)
        for _type in [*CameraType.ALL_TYPE, 'ALL']:
            cur_col_index += 1
            work_sheet.cell(cur_row_index, cur_col_index, count_data[_country][_type])
        cur_row_index += 1
    # 创建用例页面
    for _type in CameraType.ALL_TYPE:
        if _type in [CameraType.TYPE_SECTION_END, CameraType.TYPE_DANGER_ZONE_END]:
            continue
        sheet_index += 1
        wb.create_sheet(index=sheet_index, title=_type)
        work_sheet = wb.worksheets[sheet_index]
        cur_row_index = 1
        cur_col_index = 1
        # 设置第一行标题
        for attr in SafeCamera.AllAttr:
            work_sheet.cell(cur_row_index, cur_col_index, attr)
            cur_col_index += 1
        cur_row_index += 1
        for _country in data.keys():
            if data[_country].get(_type):
                for _data in data[_country].get(_type):
                    cur_col_index = 1
                    _data.set('ISO_COUNTRY_CODE', os.path.basename(_country))
                    _data = SafeCamera(_data)
                    for attr in SafeCamera.AllAttr:
                        work_sheet.cell(cur_row_index, cur_col_index, getattr(_data, attr))
                        cur_col_index += 1
                    cur_row_index += 1
    wb.save(save_name)


def main(camera_datas: list, excel_file: str):
    # make case dir
    time_now = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
    work_dir = os.path.join(".", f"{excel_file.split('.')[0]}_{time_now}")
    if not os.path.exists(os.path.join(".", work_dir)):
        os.mkdir(os.path.join(".", work_dir))
        LogUtils.info("Mkdir: %s" % work_dir)

    # 获取所有国家的文件夹路径
    all_country_folder = []
    for camera_data in camera_datas:
        all_country_folder.extend(
            return_all_folder_without_tag(aim_path=camera_data, without_tag=['DEFINITIONS', 'XSD']))
    # 获取所有文件夹-解压缩-获取所有xml-获取所有节点-过滤节点-保存-处理节点-输出
    count = {}
    all_node = {}
    for folder in all_country_folder:
        # 解压缩
        unzip_all_xml(folder)
    for folder in all_country_folder:
        # 读取文件夹下每个文件
        xml_paths = PathUtils.get_path_all_xml(folder)
        folder_node = []
        # 将每个文件的数据拼接到一起
        for xml_path in xml_paths:
            folder_node += return_all_node_in_xml(xml_path)
        # 记录这个国家的视像头数量
        count[os.path.basename(folder)] = {}
        count[os.path.basename(folder)]['ALL'] = len(folder_node)
        _folder_type_data = {}
        # 将摄像头分类放
        for pre_node in folder_node:
            pre_type = return_camera_type(pre_node)
            if pre_type in CameraType.ALL_TYPE:
                if _folder_type_data.get(pre_type) is not None:
                    _folder_type_data[pre_type].append(pre_node)
                else:
                    _folder_type_data[pre_type] = [pre_node]
            else:
                LogUtils.warning('Camera Type :%s not match!' % pre_type)
        # 记录这个国家各类摄像头数量
        for _type in CameraType.ALL_TYPE:
            if _folder_type_data.get(_type) is not None:
                count[os.path.basename(folder)][_type] = len(_folder_type_data[_type])
            else:
                count[os.path.basename(folder)][_type] = ''
        # 过滤数量
        _data_pool = {}
        for _camera_type in _folder_type_data.keys():
            # 对section end & danger end 特别处理
            if _camera_type != CameraType.TYPE_SECTION_END and _camera_type != CameraType.TYPE_DANGER_ZONE_END:
                _camera_type_data = _folder_type_data[_camera_type]
                if len(_camera_type_data) > 10:
                    pool = random.sample(_camera_type_data, 10)
                else:
                    pool = _camera_type_data
                _data_pool[_camera_type] = pool
        # 将section end数据添加到 section start中
        if _data_pool.get(CameraType.TYPE_SECTION_START) is not None:
            _data_pool[CameraType.TYPE_SECTION_START].extend(
                get_section_end_node(_data_pool.get(CameraType.TYPE_SECTION_START),
                                     _folder_type_data.get(CameraType.TYPE_SECTION_END)))
        # 将danger end数据添加到 danger start中
        if _data_pool.get(CameraType.TYPE_DANGER_ZONE_START) is not None:
            _data_pool[CameraType.TYPE_DANGER_ZONE_START].extend(
                get_section_end_node(_data_pool.get(CameraType.TYPE_DANGER_ZONE_START),
                                     _folder_type_data.get(CameraType.TYPE_DANGER_ZONE_END)))
        all_node[folder] = _data_pool
    ObjUtils.save_obj(all_node, work_dir, "all_node")
    ObjUtils.save_obj(count, work_dir, "count")
    # 这里开始解析和写excel
    save_path = os.path.join(work_dir, excel_file)
    write_excel(count, all_node, save_path)

    shutil.copy(os.path.join(".", "QueryDataCase.log"), f"./{work_dir}")


def debug(debug_dir: str):
    with open(f'{debug_dir}/all_node.pkl', 'rb') as f:
        all_node = pickle.load(f)
    with open(f'{debug_dir}/count.pkl', 'rb') as f:
        count = pickle.load(f)
    # 这里开始解析和写excel
    save_path = os.path.join(f"{debug_dir}", "debug.xlsx")
    write_excel(count, all_node, save_path)


if __name__ == '__main__':
    ev_data_path = ['E:\CameraData\Safety Cameras Malaysia S231_H0',
                    'E:\CameraData\Safety Cameras Singapore S231_H0',
                    'E:\CameraData\Safety Cameras Thailand S231_H0',
                    'E:\CameraData\Safety Cameras Vietnam S231_H0']
    excel_file_name = 'HERE_THA_S231F0.xlsx'
    main(ev_data_path, excel_file_name)
    # debug('D:\python\project\AstrobTestScript\QueryXmlSafeCamera\HERE_EU_S231H0_20240307_161131')
