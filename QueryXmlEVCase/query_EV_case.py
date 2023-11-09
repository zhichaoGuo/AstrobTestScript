import os.path
import random
import shutil
import threading
import time
from multiprocessing import pool
import xml.etree.ElementTree as ET
import openpyxl

from DataInfo import EVInfo
from Utils import PathUtils, ZipUtils, LogUtils

EV_CategoryId = ['700-7600-0322', '700-7600-0325', '700-7600-0444']


def unzip_all_xml(path: str):
    """
    使用进程池多进程的执行解压缩任务，pool.Pool应根据执行计算机的cup核心数进行设定
    目前主要受限于硬盘的读写性能
    :param path:
    :return:
    """
    start_time_stamp = time.time()
    LogUtils.info('开始解压缩！')
    p = pool.Pool(8)
    data_version_list = PathUtils.get_path_all_dir(path)
    if not data_version_list:
        LogUtils.warning('工作目录无版本路径')
        return False
    for data_version in data_version_list:
        data_country_list = PathUtils.get_path_all_dir(data_version)
        if not data_country_list:
            LogUtils.warning('[%s]版本目录无国家路径' % data_version)
            continue
        for data_country in data_country_list:
            country_zip = PathUtils.get_path_all_xml_gz(data_country)
            if not country_zip:
                LogUtils.warning('[%s]版本 [%s]国家 目录下无压缩文件' % (data_version, data_country))
                continue
            for zip_path in country_zip:
                p.apply_async(func=ZipUtils.unzip_gz_and_delete, args=(zip_path,))
    p.close()
    p.join()
    LogUtils.info('解压完成，用时：%s' % (time.time() - start_time_stamp))
    return True


def query_node_from_xml(xml_path: str) -> list:
    """
    从xml中找到符合要求的Place节点，将其转为EVInfo类型并传出为列表
    :param xml_path:
    :return:
    """
    tag = EVInfo.tag
    r = []
    f = open(xml_path, encoding='utf-8')
    tree = ET.parse(f)
    root = tree.getroot()
    # tag = '{http://places.maps.here.com/pds}'
    line_index = 1
    for node_place in root.findall(f'{tag}Place'):
        line_index += 1
        for node_category in node_place.find(f'{tag}Content').find(f'{tag}Base').find(f'{tag}CategoryList').findall(
                f'{tag}Category'):
            if node_category.get('categorySystem') != 'navteq-lcms':
                continue
            if node_category.find(f'{tag}CategoryId').text in EV_CategoryId:
                cur_ev = EVInfo(node_place)
                cur_ev.load_xml_path(xml_path)
                cur_ev.load_line_index(line_index)
                r.append(cur_ev)
    f.close()
    LogUtils.debug('[解析xml完成-> %s 个目标] PATH：%s' % (len(r), xml_path))
    return r


def filter_country_info(country_info: list) -> list:
    """
    在此处过滤国家内节点的信息，目前的原则是：
    1.国家内node数量少于20的不处理；
    2.以AdminLevel4Id属性区分城市，城市内node数量少于10的不处理；
    3.城市内node数量大于10的随机取10个；
    :param country_info: 一个EVInfo的列表
    :return: 一个EVInfo的列表
    """
    filter_info = []
    admin_level_4_id_dict = {}
    if not country_info:
        LogUtils.warning('[跳过过滤]国家目标数量为0')
        return country_info
    if len(country_info) < 20:
        LogUtils.info('[%s 跳过过滤]国家目标数量少于20' % country_info[0].CountryCode)
        return country_info
    for node in country_info:
        if node.AdminLevel4Id not in admin_level_4_id_dict.keys():
            admin_level_4_id_dict[node.AdminLevel4Id] = [node]
        else:
            admin_level_4_id_dict[node.AdminLevel4Id].append(node)
    for node in admin_level_4_id_dict.keys():
        if len(admin_level_4_id_dict[node]) < 10:
            LogUtils.debug('[%s 跳过过滤]城市目标少于10' % node)
            filter_info.extend(admin_level_4_id_dict[node])
        else:
            LogUtils.debug('[%s 经过过滤]筛选10/%s' % (node, len(admin_level_4_id_dict[node])))
            filter_info.extend(random.sample(admin_level_4_id_dict[node], 10))
    return filter_info


def write_excel(save_path: str, data: list, country_count: dict,case_path:str):
    """
    将数据写表格
    :param case_path:
    :param country_count:
    :param save_path:
    :param data:
    :return:
    """
    # 新建工作簿与工作表
    LogUtils.info('[开始构建表格]')
    wb = openpyxl.Workbook()
    # 创建总览
    wb.create_sheet(index=0, title='country count')
    work_sheet = wb.worksheets[0]
    cur_row_index = 1
    cur_col_index = 1
    for key in country_count.keys():
        work_sheet.cell(cur_row_index, cur_col_index, key)
        work_sheet.cell(cur_row_index, cur_col_index + 1, country_count[key])
        cur_row_index += 1
    # 创建用例页面
    wb.create_sheet(index=1, title=country_count["version"])
    work_sheet = wb.worksheets[1]
    cur_row_index = 1
    cur_col_index = 1
    # 设置第一行标题
    for attr in EVInfo.AllAttr:
        work_sheet.cell(cur_row_index, cur_col_index, attr)
        cur_col_index += 1
    cur_row_index += 1
    for pro_data in data:
        cur_col_index = 1
        for attr in EVInfo.AllAttr:
            work_sheet.cell(cur_row_index, cur_col_index, getattr(pro_data, attr))
            cur_col_index += 1
        cur_row_index += 1
    # 保存工作表
    version_name = os.path.basename(save_path)
    excel_name = f'{case_path}/{version_name}_EVCase.xlsx'
    wb.save(excel_name)
    LogUtils.info('[表格构建结束]')
    return excel_name


def query_all_country_node_and_write_excel(version_path: str, case_path: str,country: list = None):
    """
    查询版本路径下的所有国家的节点数据并过滤，在版本路径下保存excel
    :param case_path:
    :param country: 选定的国家列表，如果为None则为全部
    :param version_path:
    :return:
    """
    start_time = time.time()
    LogUtils.info('[开始处理 %s 版本数据]' % os.path.basename(version_path))
    version_info = []
    all_country_path = PathUtils.get_path_all_dir(version_path)
    query_country = []
    country_count = {"version": os.path.basename(version_path)}
    if country is not None:
        for cur_country in all_country_path:
            if os.path.basename(cur_country) in country:
                query_country.append(cur_country)
    else:
        query_country = all_country_path
    for country_path in query_country:
        country_info = []
        node = []
        node_number = 0
        for xml_path in PathUtils.get_path_all_xml(country_path):
            node = query_node_from_xml(xml_path)
            if node:
                # 添加数据
                node_number += len(node)
                country_info.extend(node)
        # 在此统计国家信息
        country_name = os.path.basename(country_path)
        country_count[country_name] = node_number
        # 在此处理整个国家的数据
        LogUtils.info('[开始过滤 %s 国家数据]' % os.path.basename(country_path))
        new_country_info = filter_country_info(country_info)
        LogUtils.debug(
            '[%s 过滤结果 %s/%s]' % (os.path.basename(country_path), len(new_country_info), len(country_info)))
        LogUtils.info('[过滤 %s 国家数据结束]' % os.path.basename(country_path))
        version_info.extend(new_country_info)
    # 所有国家数据处理完开始生成一张表格
    write_excel(version_path, version_info, country_count,case_path)
    LogUtils.info('[处理 %s 版本数据结束，用时：%s]' % (os.path.basename(version_path), time.time() - start_time))


def main(data_path: str, version: list = None, country: list = None):
    # 将所有的gz文件解压缩为xml文件
    unzip_all_xml(data_path)
    all_threading = []
    all_version = PathUtils.get_path_all_dir(data_path)
    query_version = []
    # 过滤选定的版本
    if version is not None:
        for cur_version in all_version:
            if os.path.basename(cur_version) in version:
                query_version.append(cur_version)
    else:
        query_version = all_version
    # 创建结果目录
    case_dir = os.path.join(os.path.dirname(data_path),'TestCase')
    if os.path.exists(case_dir):
        shutil.rmtree(case_dir)
    os.mkdir(case_dir)

    # 起多线程执行查询过滤保存操作
    for version_path in query_version:
        all_threading.append(
            threading.Thread(target=query_all_country_node_and_write_excel, args=(version_path, case_dir,country,)))
    for t in all_threading:
        t.start()
    for t in all_threading:
        t.join()
    merge_excel(case_dir)

def merge_excel(excel_path: str):
    all_excel_path = PathUtils.get_path_all_xlsx(excel_path)
    wb = openpyxl.Workbook()
    # 创建总览
    work_sheet_index = 0
    wb.create_sheet(index=work_sheet_index, title='country count')
    work_sheet = wb.worksheets[work_sheet_index]
    all_country_count = {}
    for per_excel_path in all_excel_path:
        per_excel_sheet = openpyxl.load_workbook(per_excel_path)["country count"]
        version = openpyxl.load_workbook(per_excel_path)["country count"].cell(1, 2).value
        for row in per_excel_sheet.iter_rows():
            line = [cell.value for cell in row]
            if line[0] in all_country_count.keys():
                all_country_count[line[0]][version] = line[1]
            else:
                all_country_count[line[0]] = {version: line[1]}
    cur_col_index = 1
    cur_row_index = 1
    work_sheet.cell(cur_row_index, cur_col_index, "country")
    for version in all_country_count["version"].keys():
        cur_col_index += 1
        work_sheet.cell(cur_row_index, cur_col_index, version)
    for country in all_country_count.keys():
        if country != "country count":
            cur_col_index = 1
            cur_row_index += 1
            work_sheet.cell(cur_row_index,cur_col_index,country)
            for version in all_country_count["version"].keys():
                cur_col_index += 1
                # count = all_country_count[country].get(version)
                work_sheet.cell(cur_row_index,cur_col_index,all_country_count[country].get(version))


    # 合并用例
    for per_excel_path in all_excel_path:
        work_sheet_index += 1
        sheet_title = os.path.basename(per_excel_path).split('.')[0]
        new_sheet = wb.create_sheet(index=work_sheet_index, title=sheet_title)
        version = openpyxl.load_workbook(per_excel_path)["country count"].cell(1, 2).value
        per_excel_sheet = openpyxl.load_workbook(per_excel_path)[version]
        for row in per_excel_sheet.iter_rows():
            new_sheet.append([cell.value for cell in row])
    merge_name = os.path.join(excel_path, 'merge.xlsx')
    wb.save(merge_name)


if __name__ == '__main__':
    # main(data_path='E:\\PLACES_XML_PREMIUM_23118_EUW_DPLAP\\WEU_23118')
    # main(data_path='E:\\PLACES_XML_PREMIUM_231F1_WVM_DPLAP\\MEA_231F0')
    # main(data_path='E:\\test')
    main(data_path='E:\\PLACES_XML_PREMIUM_23118_EUW_DPLAP\\WEU_23118',
         version=['TQS2'],
         country=['IRL'])
    # merge_excel('E:\PLACES_XML_PREMIUM_231F1_WVM_DPLAP\TestCase')
