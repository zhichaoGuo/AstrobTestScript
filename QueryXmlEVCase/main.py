import os
import pickle
import random
import shutil
import time

import xml.etree.ElementTree as ET

import openpyxl

from QueryXmlEVCase.EVChargePoint import EVChargePoint
from QueryXmlEVCase.QueryScript import QueryISOCountryCode
from Utils import LogUtils, PathUtils, ZipUtils, SqlServer, load_config, ObjUtils


def return_all_folder_without_tag(aim_path: str, without_tag: list):
    """
    返回文件夹中除了在without中的文件夹的绝对路径
    """
    r = []
    for path_file in os.listdir(aim_path):
        item_path = os.path.join(aim_path, path_file)
        if os.path.isdir(item_path) and os.path.basename(item_path) not in without_tag:
            if len(os.path.basename(item_path)) != 3:
                print('文件路径长度不等于3请关注：%s' % item_path)
            else:
                r.append(item_path)
    return r


def unzip_all_xml(path: str):
    """
    解压缩路径中的压缩包
    """
    zip_paths = PathUtils.get_path_all_xml_gz(path)
    if not zip_paths:
        LogUtils.warning('[%s]目录下无压缩文件' % path)
    for zip_path in zip_paths:
        ZipUtils.unzip_gz_and_delete(zip_path)


def return_all_node_in_xml(x_path: str):
    f = open(x_path, encoding='utf-8')
    tree = ET.parse(f)
    root = tree.getroot()
    # all_node = []
    #
    # for node_poi in random.sample(root.findall(f'POI'), 10):
    #     all_node.append(EVChargePoint(node_poi))
    return root.findall(f'POI')


def write_excel(count_data: dict, write_data: dict, save_name: str):
    wb = openpyxl.Workbook()
    # 创建总览
    sheet_index = 0
    wb.create_sheet(index=sheet_index, title='country count')
    work_sheet = wb.worksheets[0]
    cur_row_index = 1
    cur_col_index = 1
    for key in count_data.keys():
        work_sheet.cell(cur_row_index, cur_col_index, key)
        work_sheet.cell(cur_row_index, cur_col_index + 1, count_data[key])
        cur_row_index += 1
    # 创建用例页面
    sheet_index += 1
    wb.create_sheet(index=sheet_index, title='data case')
    work_sheet = wb.worksheets[sheet_index]
    cur_row_index = 1
    cur_col_index = 1
    # 设置第一行标题
    for attr in EVChargePoint.AllAttr:
        work_sheet.cell(cur_row_index, cur_col_index, attr)
        cur_col_index += 1
    cur_row_index += 1

    for _, pro_data in write_data.items():
        for _data in pro_data:
            cur_col_index = 1
            for attr in EVChargePoint.AllAttr:
                work_sheet.cell(cur_row_index, cur_col_index, getattr(_data, attr))
                cur_col_index += 1
            cur_row_index += 1
    # for pro_iso_code, pro_data in write_data.items():
    #     sheet_index += 1
    #     # 创建用例页面
    #     wb.create_sheet(index=sheet_index, title=pro_iso_code)
    #     work_sheet = wb.worksheets[sheet_index]
    #     cur_row_index = 1
    #     cur_col_index = 1
    #     # 设置第一行标题
    #     for attr in EVChargePoint.AllAttr:
    #         work_sheet.cell(cur_row_index, cur_col_index, attr)
    #         cur_col_index += 1
    #     cur_row_index += 1
    #     for _data in pro_data:
    #         cur_col_index = 1
    #         for attr in EVChargePoint.AllAttr:
    #             work_sheet.cell(cur_row_index, cur_col_index, getattr(_data, attr))
    #             cur_col_index += 1
    #         cur_row_index += 1
    # 保存工作表
    # version_name = os.path.basename(save_path)
    # excel_name = save_name
    # excel_name = os.path.join(os.path.dirname(__file__), 'case', save_name)
    wb.save(save_name)


class DbManager:
    def __init__(self, db_list: list):
        self.db_obj_dict = {}
        self.country_code_dict = {}
        for db in db_list:
            self.db_obj_dict[db] = SqlServer(**load_config().get("sqlserver"), database=db)
            self.country_code_dict[db] = self.get_all_country_code(self.db_obj_dict[db])

    def get_all_country_code(self, db):
        obj = db.query(QueryISOCountryCode())
        res = []
        for o in obj:
            res.append(o.ISO_COUNTRY_CODE)
        print('get db all country code:%s' % res)
        return res

    def get_db_by_iso_cc(self, iso_country_code: str):
        for db_name in self.country_code_dict.keys():
            if iso_country_code in self.country_code_dict[db_name]:
                return self.db_obj_dict[db_name]
        print('can not found iso country code in dbm:%s' % iso_country_code)
        return None

    def close(self):
        for db_name in self.db_obj_dict.keys():
            self.db_obj_dict[db_name].close()


def main(ev_data: str, excel_file: str, sql_db: list,pre_country_count:int=10):
    time_now = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
    work_dir = os.path.join(".", f"{excel_file.split('.')[0]}_{time_now}")
    if not os.path.exists(os.path.join(".", work_dir)):
        os.mkdir(os.path.join(".", work_dir))
        LogUtils.info("Mkdir: %s" % work_dir)
    dbm = DbManager(sql_db)
    # 获取所有国家的文件夹路径
    all_country_folder = return_all_folder_without_tag(aim_path=ev_data, without_tag=['Reference'])
    count = {}
    data = {}
    # 获取所有文件夹-解压缩-获取所有xml-获取所有节点-过滤节点-保存-处理节点-输出
    all_node = {}
    pool_node = {}
    for folder in all_country_folder:
        # 解压缩
        unzip_all_xml(folder)
        xml_paths = PathUtils.get_path_all_xml(folder)
        folder_node = []
        for xml_path in xml_paths:
            folder_node += return_all_node_in_xml(xml_path)
        count[os.path.basename(folder)] = len(folder_node)
        all_node[folder] = folder_node

        if len(folder_node) > pre_country_count:
            pool = random.sample(folder_node, pre_country_count)
        else:
            pool = folder_node
        pool_node[folder] = pool
    # ObjUtils.save_obj(all_node, work_dir, "all_node")
    for folder in pool_node.keys():
        for node in pool_node[folder]:
            node.set('uuid', all_node[folder].index(node) + 1)
    ObjUtils.save_obj(pool_node, work_dir, "pool_node")
    ObjUtils.save_obj(count, work_dir, "count")
    ObjUtils.save_obj(sql_db, work_dir, "sql_db")
    for folder in pool_node.keys():
        folder_obj = []
        for node in pool_node[folder]:
            ev_obj = EVChargePoint(node, dbm)
            ev_obj.add_index(all_node[folder].index(node) + 1)
            folder_obj.append(ev_obj)
        data[os.path.basename(folder)] = folder_obj
    save_path = os.path.join(work_dir, excel_file)
    write_excel(count, data, save_path)
    dbm.close()


def debug(debug_dir: str):
    # 从文件中加载实例
    # with open(f'{debug_dir}/all_node.pkl', 'rb') as f:
    #     all_node = pickle.load(f)
    with open(f'{debug_dir}/pool_node.pkl', 'rb') as f:
        pool_node = pickle.load(f)
    with open(f'{debug_dir}/count.pkl', 'rb') as f:
        count = pickle.load(f)
    with open(f'{debug_dir}/sql_db.pkl', 'rb') as f:
        sql_db = pickle.load(f)
    dbm = DbManager(sql_db)
    data = {}
    for folder in pool_node.keys():
        folder_obj = []
        for node in pool_node[folder]:
            ev_obj = EVChargePoint(node, dbm)
            ev_obj.add_index(node.get('uuid'))
            folder_obj.append(ev_obj)
        data[os.path.basename(folder)] = folder_obj
    save_path = os.path.join(f"{debug_dir}", "debug.xlsx")
    write_excel(count, data, save_path)
    dbm.close()
    # shutil.move(os.path.join(".", "QueryDateCase.log"), f"./{debug_dir}_debug")


if __name__ == '__main__':
    country_count=10
    # APAC=============================
    # ev_data_path = 'E:\EVdata\HERE EV Charge Points Static Asia Pacific S231_H0'
    # sql_db_name = ['HERE_APAC_S231R4']
    # excel_file_name = 'HERE_APAC_231H0.xlsx'
    # MEA==============================
    # ev_data_path = 'E:\EVdata\HERE EV Charge Points Static MEA S231_H0'
    # sql_db_name = ['HERE_MEA_S231R4']
    # excel_file_name = 'HERE_MEA_231H0.xlsx'
    # EU===============================
    # ev_data_path = 'E:\EVdata\HERE EV Charge Points Static South America S231_H0'
    sql_db_name = ['HERE_SAM_S231R4']
    excel_file_name = 'HERE_SAM_231H0.xlsx'
    # IND==============================
    ev_data_path = 'E:\EVdata\HERE EV Charge Points Static India S231_H0'
    sql_db_name = ['HERE_IND_S231R4']
    excel_file_name = 'HERE_IND_231H0.xlsx'
    country_count = 30
    # sql_db_name = []

    main(ev_data_path, excel_file_name, sql_db_name,pre_country_count=country_count)
    # debug_path = "D:\python\project\AstrobTestScript\QueryXmlEVCase\HERE_APAC_231H0_20240221_154816"
    # debug(debug_path)
