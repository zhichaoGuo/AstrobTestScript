import os.path

import openpyxl
from openpyxl.styles import PatternFill

from Utils import PathUtils, LogUtils

LanguageFileDir = os.path.join(os.path.dirname(__file__), 'Language')


def load_dat(dat_file_path) -> dict:
    """

    """
    with open(dat_file_path, 'r', encoding='utf-8') as f:
        f_header = ''
        f_body = {}
        first_line = True
        for line in f:
            if first_line:  # 首行为65001
                f_header = line
                first_line = False
            else:
                try:
                    f_index, f_line = line.split("=")
                    f_body[f_index] = f_line.split(',"')[1].replace('"', '')
                except Exception as err:
                    LogUtils.error('[解析dat文件行发生错误]：%s' % err)
                    LogUtils.error('错误文件：%s ,错误行：%s' % (dat_file_path, line))

    return f_body


def write_excel(all_key: list, all_data: dict):
    fill_exist = PatternFill('solid', fgColor="76923C")
    fill_not_exist = PatternFill('solid', fgColor="D7E3BC")
    wb = openpyxl.Workbook()
    sheet_index = 0
    wb.create_sheet(index=sheet_index, title='country_data')
    cur_sheet = wb.worksheets[sheet_index]
    cur_row_index = 1
    cur_col_index = 2
    # 设置第一行标题
    for lang in all_data.keys():
        cur_sheet.cell(cur_row_index, cur_col_index, lang)
        cur_col_index += 1
    cur_row_index += 1

    for key in all_key:
        cur_col_index = 1
        cur_sheet.cell(cur_row_index, cur_col_index, key)
        cur_col_index += 1
        for data in all_data.values():
            if data.get(key):
                _fill = fill_exist
            else:
                _fill = fill_not_exist
            cur_sheet.cell(cur_row_index, cur_col_index, data.get(key)).fill = _fill
            cur_col_index += 1
        cur_row_index += 1
    db_name = '语言对比'
    wb.save(f'./{db_name}.xlsx')


def main():
    all_lang_dat_path = PathUtils.get_path_all_dat(LanguageFileDir)
    all_lang_obj = {}
    all_key = []
    for dat_path in all_lang_dat_path:
        lang_data = load_dat(dat_path)
        all_lang_obj[os.path.basename(dat_path)] = lang_data
        for key in lang_data.keys():
            if key not in all_key:
                all_key.append(key)
    write_excel(all_key, all_lang_obj)


if __name__ == '__main__':
    main()
